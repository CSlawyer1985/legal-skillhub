#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
税务合规自查引擎 (Tax Compliance Self-Check Engine)
====================================================
读取企业税务事实 JSON，依 scripts/rules.json 中的规则基线做自动化风险扫描，
输出：1) 控制台 Markdown 摘要；2) Excel 自查报告（风险清单 / 自查汇总 / 待确认项）。

零业务依赖，仅 openpyxl（Excel 输出时使用）。纯 Python 标准库即可跑控制台摘要。

用法:
    python run_check.py --input sample_data/sample_input.json --out 税务合规自查报告.xlsx
    python run_check.py -i facts.json -o report.xlsx --unit wan   # 金额单位仅影响展示

退出码: 0=正常(含风险); 2=输入/规则错误(已给出中文提示)。
"""
import argparse
import json
import os
import sys
from datetime import datetime

SEVERITY_WEIGHT = {"高": 3, "中": 2, "低": 1}
SEVERITY_ORDER = {"高": 0, "中": 1, "低": 2}
_MISSING = object()


def die(msg, code=2):
    print("✗ " + msg, file=sys.stderr)
    sys.exit(code)


def load_json(path, what):
    if not os.path.isfile(path):
        die(f"{what}文件不存在或路径错误：{path}\n  请检查路径、文件名与读写权限。")
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except json.JSONDecodeError as e:
        die(f"{what}JSON 解析失败（第 {e.lineno} 行第 {e.colno} 列）：{e.msg}\n"
            f"  常见原因：引号未闭合、逗号多余/缺失、误用中文标点、括号不配对。")


def get_path(obj, path):
    cur = obj
    for seg in path.split("."):
        if isinstance(cur, dict) and seg in cur:
            cur = cur[seg]
        else:
            return _MISSING
    return cur


def is_missing(v):
    return v is _MISSING


def is_filled(v):
    if v is _MISSING:
        return False
    if isinstance(v, bool):
        return v
    if isinstance(v, (list, dict, str)):
        return len(v) > 0
    if isinstance(v, (int, float)):
        return v != 0
    return bool(v)


def to_num(v):
    if isinstance(v, bool):
        return 1.0 if v else 0.0
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        return float(v.replace(",", "").replace("，", ""))
    return 0.0


def eval_condition(cond, facts):
    """返回 (triggered, value) —— triggered: True=风险命中, False=通过, None=无法评估(待确认)。"""
    t = cond.get("type")
    if t == "flag_true":
        v = get_path(facts, cond["path"])
        if is_missing(v):
            return None, v
        return is_filled(v), v
    if t == "flag_false":
        v = get_path(facts, cond["path"])
        if is_missing(v):
            return None, v
        return (not is_filled(v)), v
    if t == "missing":
        v = get_path(facts, cond["path"])
        return (is_missing(v) or not is_filled(v)), v
    if t in ("gt", "lt", "ge", "le"):
        v = get_path(facts, cond["path"])
        if is_missing(v):
            return None, v
        a = to_num(v)
        b = float(cond["value"])
        if t == "gt":
            return a > b, v
        if t == "lt":
            return a < b, v
        if t == "ge":
            return a >= b, v
        return a <= b, v
    if t == "diff_pct":
        l = get_path(facts, cond["left_path"])
        r = get_path(facts, cond["right_path"])
        if is_missing(l) or is_missing(r):
            return None, (l, r)
        denom = max(abs(to_num(l)), abs(to_num(r)), 1.0)
        pct = abs(to_num(l) - to_num(r)) / denom * 100.0
        return pct > float(cond["max_pct"]), (to_num(l), to_num(r), round(pct, 2))
    if t == "meal_limit":
        amt = get_path(facts, cond["path"])
        base = get_path(facts, cond["base_path"])
        if is_missing(amt) or is_missing(base):
            return None, (amt, base)
        # 可扣除 = min(发生额×60%, 营收×5‰)；超过则未调增风险
        return to_num(amt) * 0.6 > to_num(base) * 0.005, (to_num(amt), to_num(base))
    if t == "deduct_limit":
        amt = get_path(facts, cond["path"])
        base = get_path(facts, cond["base_path"])
        if is_missing(amt) or is_missing(base):
            return None, (amt, base)
        return to_num(amt) > to_num(base) * float(cond["rate"]), (to_num(amt), to_num(base))
    if t in ("all", "any"):
        subs = [eval_condition(c, facts) for c in cond.get("conds", [])]
        if any(s[0] is None for s in subs):
            return None, subs
        if t == "all":
            return all(s[0] for s in subs), subs
        return any(s[0] for s in subs), subs
    die(f"规则引擎遇到未知条件类型：{t}")


def fmt_value(v):
    if is_missing(v):
        return "（未提供）"
    if isinstance(v, bool):
        return "是" if v else "否"
    if isinstance(v, (list,)):
        return "、".join(str(x) for x in v) if v else "（空）"
    if isinstance(v, (int, float)):
        return f"{v:,.2f}"
    return str(v)


def build_report(input_data, rules_doc, unit="yuan"):
    company = input_data.get("company", {})
    rules = rules_doc.get("rules", [])

    results = []
    for rule in rules:
        # 规则路径以 facts.xxx 开头，故以整份输入为根对象解释执行
        triggered, val = eval_condition(rule["condition"], input_data)
        results.append({
            "id": rule["id"],
            "category": rule["category"],
            "name": rule["name"],
            "severity": rule["severity"],
            "regulation": rule.get("regulation", ""),
            "remediation": rule.get("remediation", ""),
            "triggered": triggered,
            "value": val,
        })

    triggered_list = [r for r in results if r["triggered"] is True]
    passed_list = [r for r in results if r["triggered"] is False]
    pending_list = [r for r in results if r["triggered"] is None]

    # 合规评分：仅基于已评估规则按严重度加权
    eval_weight = sum(SEVERITY_WEIGHT[r["severity"]] for r in results if r["triggered"] is not None)
    trig_weight = sum(SEVERITY_WEIGHT[r["severity"]] for r in triggered_list)
    score = round(100 * (eval_weight - trig_weight) / eval_weight, 1) if eval_weight else 100.0

    sev_counts = {"高": 0, "中": 0, "低": 0}
    for r in triggered_list:
        sev_counts[r["severity"]] += 1

    triggered_list.sort(key=lambda r: (SEVERITY_ORDER[r["severity"]], r["id"]))

    return {
        "company": company,
        "results": results,
        "triggered": triggered_list,
        "passed": passed_list,
        "pending": pending_list,
        "score": score,
        "sev_counts": sev_counts,
        "unit": unit,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def render_markdown(rep):
    c = rep["company"]
    name = c.get("name", "（未命名企业）")
    period = c.get("period", "—")
    unit_label = "万元" if rep["unit"] == "wan" else "元"
    lines = []
    lines.append(f"# 税务合规自查报告 · {name}")
    lines.append("")
    lines.append(f"- 所属期：{period}　|　生成时间：{rep['generated_at']}　|　金额单位：{unit_label}")
    lines.append(f"- **合规评分：{rep['score']} 分**（满分100，越高越合规）")
    sc = rep["sev_counts"]
    lines.append(f"- **命中风险：{len(rep['triggered'])} 项**（高 {sc['高']} / 中 {sc['中']} / 低 {sc['低']}）"
                 f"　|　通过 {len(rep['passed'])} 项　|　待确认 {len(rep['pending'])} 项")
    lines.append("")
    if rep["triggered"]:
        lines.append("## 一、风险清单（按严重度排序）")
        lines.append("")
        lines.append("| 严重度 | 编号 | 类别 | 风险点 | 法规依据 | 整改建议 |")
        lines.append("| --- | --- | --- | --- | --- | --- |")
        for r in rep["triggered"]:
            lines.append(f"| {r['severity']} | {r['id']} | {r['category']} | {r['name']} | "
                         f"{r['regulation']} | {r['remediation']} |")
        lines.append("")
    if rep["pending"]:
        lines.append("## 二、待确认项（数据缺失，无法评估）")
        lines.append("")
        lines.append("| 编号 | 类别 | 风险点 | 需补充的数据 |")
        lines.append("| --- | --- | --- | --- |")
        for r in rep["pending"]:
            lines.append(f"| {r['id']} | {r['category']} | {r['name']} | 见 rules.json 的 {r['id']} 条件路径 |")
        lines.append("")
    lines.append("## 三、结论与建议")
    lines.append("")
    if not rep["triggered"]:
        lines.append("- 本次自查范围内未发现明确合规风险点；待确认项请补充数据后复评。")
    else:
        top = "、".join(f"{r['id']}({r['severity']})" for r in rep["triggered"][:3])
        lines.append(f"- 优先处置高风险项（{top}），逐项落实整改建议并留存证据链。")
        lines.append("- 中低风险纳入日常税务内控，建立申报日历与凭证复核清单。")
        lines.append("- 本报告为辅助自查工具，重大涉税事项以最新法规与主管税务机关意见为准，建议由税务师复核。")
    return "\n".join(lines)


def write_excel(rep, out_path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("⚠ 未安装 openpyxl，跳过 Excel 输出。仅控制台 Markdown 摘要可用。\n"
              "  安装：pip install openpyxl", file=sys.stderr)
        return False

    wb = Workbook()
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    sev_fill = {"高": PatternFill("solid", fgColor="F4CCCC"),
                "中": PatternFill("solid", fgColor="FCE5CD"),
                "低": PatternFill("solid", fgColor="FFF2CC")}
    wrap = Alignment(wrap_text=True, vertical="top")

    # Sheet 1: 风险清单
    ws = wb.active
    ws.title = "风险清单"
    headers = ["严重度", "编号", "类别", "风险点", "当前值", "法规依据", "整改建议"]
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for r in rep["triggered"]:
        ws.append([r["severity"], r["id"], r["category"], r["name"],
                   fmt_value(r["value"]), r["regulation"], r["remediation"]])
        row = ws.max_row
        ws.cell(row, 1).fill = sev_fill.get(r["severity"])
        for col in range(1, len(headers) + 1):
            ws.cell(row, col).alignment = wrap
            ws.cell(row, col).border = border
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["G"].width = 48
    ws.freeze_panes = "A2"

    # Sheet 2: 自查汇总
    ws2 = wb.create_sheet("自查汇总")
    c = rep["company"]
    rows = [
        ["企业名称", c.get("name", "（未命名企业）")],
        ["纳税人类型", c.get("taxpayer_type", "—")],
        ["所属行业", c.get("industry", "—")],
        ["所属期", c.get("period", "—")],
        ["生成时间", rep["generated_at"]],
        ["金额单位", "万元" if rep["unit"] == "wan" else "元"],
        ["合规评分", rep["score"]],
        ["命中风险总数", len(rep["triggered"])],
        ["  其中高风险", rep["sev_counts"]["高"]],
        ["  其中中风险", rep["sev_counts"]["中"]],
        ["  其中低风险", rep["sev_counts"]["低"]],
        ["通过项", len(rep["passed"])],
        ["待确认项", len(rep["pending"])],
    ]
    ws2.append(["项目", "内容"])
    for col in (1, 2):
        ws2.cell(1, col).fill = header_fill
        ws2.cell(1, col).font = header_font
        ws2.cell(1, col).border = border
    for k, v in rows:
        ws2.append([k, v])
        for col in (1, 2):
            ws2.cell(ws2.max_row, col).border = border
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 40

    # Sheet 3: 待确认项
    ws3 = wb.create_sheet("待确认项")
    h3 = ["编号", "类别", "风险点", "说明"]
    ws3.append(h3)
    for col, _ in enumerate(h3, 1):
        ws3.cell(1, col).fill = header_fill
        ws3.cell(1, col).font = header_font
        ws3.cell(1, col).border = border
    if rep["pending"]:
        for r in rep["pending"]:
            ws3.append([r["id"], r["category"], r["name"], "数据缺失，需在输入 JSON 中补充对应字段后复评"])
            for col in range(1, len(h3) + 1):
                ws3.cell(ws3.max_row, col).alignment = wrap
                ws3.cell(ws3.max_row, col).border = border
    else:
        ws3.append(["—", "—", "无", "所有规则均已基于提供的数据完成评估"])
    ws3.column_dimensions["A"].width = 10
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 34
    ws3.column_dimensions["D"].width = 50

    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    wb.save(out_path)
    return True


def main():
    here = os.path.dirname(os.path.abspath(__file__))
    ap = argparse.ArgumentParser(description="企业税务合规自查引擎")
    ap.add_argument("-i", "--input", required=True, help="企业税务事实 JSON 路径")
    ap.add_argument("-o", "--out", default="税务合规自查报告.xlsx", help="Excel 报告输出路径（默认 ./税务合规自查报告.xlsx）")
    ap.add_argument("-r", "--rules", default=os.path.join(here, "rules.json"), help="规则基线 JSON 路径")
    ap.add_argument("--unit", default="yuan", choices=["yuan", "wan"], help="金额展示单位：yuan(元,默认) / wan(万元)")
    ap.add_argument("--no-excel", action="store_true", help="仅输出控制台 Markdown，不生成 Excel")
    args = ap.parse_args()

    rules_doc = load_json(args.rules, "规则基线")
    if "rules" not in rules_doc or not isinstance(rules_doc["rules"], list):
        die("规则基线格式错误：缺少 rules 数组。")
    ids = [r.get("id") for r in rules_doc["rules"]]
    if len(ids) != len(set(ids)):
        die("规则基线存在重复 id，请检查 rules.json。")

    input_data = load_json(args.input, "输入")
    if "facts" not in input_data:
        die("输入 JSON 缺少 facts 对象。请参照 sample_data/sample_input.json 组织字段。")

    rep = build_report(input_data, rules_doc, unit=args.unit)
    print(render_markdown(rep))
    print("")

    if not args.no_excel:
        if write_excel(rep, args.out):
            print(f"✓ Excel 报告已生成：{args.out}")


if __name__ == "__main__":
    main()
