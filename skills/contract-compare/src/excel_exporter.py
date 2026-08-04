"""
Excel exporter for contract comparison results.
Generates structured diff reports in .xlsx format.
"""
import os
import csv
from typing import List, Dict, Any, Optional
from datetime import datetime


class ExcelExportError(Exception):
    """Raised when Excel export fails."""
    pass


def export_diff_to_excel(diff_items: List[Dict[str, Any]], output_path: str = None) -> str:
    """Export diff items to an Excel file.

    Args:
        diff_items: List of diff items from compare_clauses
        output_path: Output file path (auto-generated if None)

    Returns:
        Path to the generated Excel file

    Raises:
        ExcelExportError: If export fails
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise ExcelExportError("openpyxl not installed")

    if output_path is None:
        os.makedirs("/tmp/contract-compare/", exist_ok=True)
        output_path = f"/tmp/contract-compare/diff_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "合同差异清单"

        # Headers
        headers = ["类型", "条款编号", "条款标题", "原内容", "新内容", "差异说明"]
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Type colors
        type_colors = {
            "new": "C6EFCE",      # Green
            "deleted": "FFC7CE",   # Red
            "modified": "FFEB9C",  # Yellow
            "same": "F2F2F2",      # Gray
        }

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        type_labels = {
            "new": "新增条款",
            "deleted": "删除条款",
            "modified": "修改条款",
            "same": "相同条款",
        }

        for row_idx, item in enumerate(diff_items, 2):
            diff_type = item.get("type", "same")
            fill_color = type_colors.get(diff_type, "FFFFFF")

            clause_a = item.get("clause_a") or {}
            clause_b = item.get("clause_b") or {}
            diff_content = item.get("diff_content") or ""

            row_data = [
                type_labels.get(diff_type, diff_type),
                clause_b.get("number") or clause_a.get("number", ""),
                clause_b.get("title") or clause_a.get("title", ""),
                clause_a.get("content", "")[:500] if clause_a else "",
                clause_b.get("content", "")[:500] if clause_b else "",
                diff_content,
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border

        # Set column widths
        col_widths = [12, 15, 20, 40, 40, 50]
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + col_idx)].width = width

        # Set row heights
        for row in range(2, len(diff_items) + 2):
            ws.row_dimensions[row].height = 60

        # Summary sheet
        ws2 = wb.create_sheet("统计摘要")
        summary = build_summary(diff_items)
        ws2["A1"] = "合同差异统计"
        ws2["A1"].font = Font(bold=True, size=14)

        ws2["A3"] = "总差异数"
        ws2["B3"] = summary["total_changes"]
        ws2["A4"] = "新增条款"
        ws2["B4"] = summary["new_count"]
        ws2["A5"] = "删除条款"
        ws2["B5"] = summary["deleted_count"]
        ws2["A6"] = "修改条款"
        ws2["B6"] = summary["modified_count"]
        ws2["A7"] = "相同条款"
        ws2["B7"] = summary["same_count"]

        for row in range(3, 8):
            ws2.row_dimensions[row].height = 25

        ws2.column_dimensions["A"].width = 20
        ws2.column_dimensions["B"].width = 15

        wb.save(output_path)
        return output_path

    except Exception as e:
        raise ExcelExportError(f"Failed to export Excel: {e}")


def build_summary(diff_items: List[Dict[str, Any]]) -> Dict[str, int]:
    """Build summary statistics from diff items.

    Args:
        diff_items: List of diff items

    Returns:
        Dict with counts
    """
    summary = {
        "total_changes": 0,
        "new_count": 0,
        "deleted_count": 0,
        "modified_count": 0,
        "same_count": 0,
    }

    for item in diff_items:
        t = item.get("type", "same")
        summary["total_changes"] += 1
        if t == "new":
            summary["new_count"] += 1
        elif t == "deleted":
            summary["deleted_count"] += 1
        elif t == "modified":
            summary["modified_count"] += 1
        else:
            summary["same_count"] += 1

    return summary


def export_multi_version_timeline(files: List[Dict[str, str]], output_path: str = None) -> str:
    """Export multi-version timeline to Excel.

    Args:
        files: List of dicts with keys: filename, version_label, date, content
        output_path: Output file path

    Returns:
        Path to the generated Excel file
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ExcelExportError("openpyxl not installed")

    if output_path is None:
        os.makedirs("/tmp/contract-compare/", exist_ok=True)
        output_path = f"/tmp/contract-compare/timeline_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "版本时间轴"

        headers = ["版本", "文件名", "版本标签", "日期"]
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, f in enumerate(files, 2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1)
            ws.cell(row=row_idx, column=2, value=f.get("filename", ""))
            ws.cell(row=row_idx, column=3, value=f.get("version_label", f"版本{row_idx-1}"))
            ws.cell(row=row_idx, column=4, value=f.get("date", ""))

        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 25

        wb.save(output_path)
        return output_path

    except Exception as e:
        raise ExcelExportError(f"Failed to export timeline Excel: {e}")


def export_risk_report(risk_items: List[Dict[str, Any]], output_path: str = None) -> str:
    """Export risk assessment report to Excel.

    Args:
        risk_items: List of risk assessments
        output_path: Output file path

    Returns:
        Path to the generated Excel file
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise ExcelExportError("openpyxl not installed")

    if output_path is None:
        os.makedirs("/tmp/contract-compare/", exist_ok=True)
        output_path = f"/tmp/contract-compare/risk_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "风险评估报告"

        headers = ["风险等级", "条款编号", "条款标题", "风险说明"]
        risk_colors = {
            "high": "FF0000",
            "medium": "FFC000",
            "low": "92D050",
        }
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        level_labels = {"high": "高风险", "medium": "中风险", "low": "低风险"}

        for row_idx, item in enumerate(risk_items, 2):
            clause = item.get("clause", {}) or {}
            risk_level = item.get("risk_level", "low")
            risk_reason = item.get("risk_reason", "")

            color_hex = risk_colors.get(risk_level, "FFFFFF")

            row_data = [
                level_labels.get(risk_level, risk_level),
                clause.get("number", ""),
                clause.get("title", ""),
                risk_reason,
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border

            # Color the risk level cell
            ws.cell(row=row_idx, column=1).fill = PatternFill(
                start_color=color_hex, end_color=color_hex, fill_type="solid"
            )

        col_widths = [12, 15, 25, 60]
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + col_idx)].width = width

        for row in range(2, len(risk_items) + 2):
            ws.row_dimensions[row].height = 40

        wb.save(output_path)
        return output_path

    except Exception as e:
        raise ExcelExportError(f"Failed to export risk report: {e}")
