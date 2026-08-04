#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""根据 analysis.json 生成劳动合同合规审查报告（HTML + Excel）。

用法:
    python build_report.py <analysis.json> [输出目录]

analysis.json 结构:
{
  "file_name": "劳动合同样本.docx",
  "review_date": "2026-07-20",
  "clause_count": 13,
  "overall_rating": "高风险",
  "overall_summary": "...",
  "local_wage": "用人单位所在地最低工资（示例·上海）：2690 元/月（2026-07 实时核对）",
  "jurisdiction": {"country": "中国", "province": "上海市", "city": "上海市"},
  "local_rules": [
    {"level": "省级", "topic": "最低工资", "value": "2740 元/月", "source": "上海市人社局", "date": "2026-01"},
    {"level": "市级", "topic": "工资支付条例", "value": "离职3个工作日内结清", "source": "《上海市企业工资支付办法》", "date": "2016"}
  ],
  "risk_cards": [ {"title": "...", "desc": "...", "level": "高"}, ... ],
  "rows": [
    {
      "index": 1,
      "quote": "原文摘要",
      "dimension": "试用期",
      "conclusion": "违规",
      "risk": "高",
      "law": "《劳动合同法》第19条",
      "suggestion": "建议改正说明",
      "compliant_example": "可直接替换的合规表述",
      "employee_advice": "站在员工立场的应对建议（可要求修改/拒签、留证、投诉/仲裁/2N 等）",
      "is_high": true
    }
  ],
  "disclaimer": "..."
}

依赖: openpyxl  (pip install openpyxl)
"""
import sys
import os
import json
import html as html_mod

RISK_BADGE = {"高": "b-high", "中": "b-mid", "低": "b-low", "合规": "b-ok"}
RISK_FILL = {
    "高": "FEE4E2", "中": "FEF0C7", "低": "E0F2FE", "合规": "D1FADF",
}
RISK_FONT = {
    "高": "D92D20", "中": "B54708", "低": "075985", "合规": "079455",
}
# 风险级别降序排序权重（高 > 中 > 低 > 合规）
RISK_ORDER = {"高": 0, "中": 1, "低": 2, "合规": 3}

# 标准免责声明：每份报告必须包含，取自 analysis.json 的 disclaimer 字段；
# 若未提供则由脚本以本默认文本填充。两条核心意思不得删减。
DEFAULT_DISCLAIMER = (
    "1. 内容性质：本审查报告及分析内容由 AI 辅助生成，仅做参考，"
    "不构成正式的法律意见或具有法律约束力的文件。\n"
    "2. 合规责任：用人单位在使用前，必须结合自身实际情况，"
    "咨询专业法律人士或人力资源顾问进行审核和修订，以确保完全符合现行法律法规。"
)

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{title}</title>
<style>
  :root {{
    --bg:#f7f8fa; --card:#fff; --ink:#1f2329; --muted:#6b7280; --line:#e5e7eb;
    --red:#d92d20; --orange:#f79009; --green:#079455; --blue:#2563eb;
  }}
  * {{ box-sizing:border-box; }}
  body {{ margin:0; font-family:-apple-system,"Segoe UI","PingFang SC","Microsoft YaHei",sans-serif;
    background:var(--bg); color:var(--ink); line-height:1.6; }}
  .wrap {{ max-width:1000px; margin:0 auto; padding:32px 24px 64px; }}
  header {{ background:linear-gradient(135deg,#2563eb,#1e40af); color:#fff; border-radius:16px;
    padding:28px 32px; margin-bottom:24px; }}
  header.danger {{ background:linear-gradient(135deg,#d92d20,#9b1c1c); }}
  header h1 {{ margin:0 0 8px; font-size:24px; }}
  .meta {{ }};
  .meta div {{ display:inline-block; margin-right:18px; font-size:13px; opacity:.95; }}
  .meta div span {{ opacity:.8; margin-right:6px; }}
  .overall {{ display:inline-flex; align-items:center; gap:8px; margin-top:16px; padding:8px 16px;
    border-radius:999px; background:rgba(255,255,255,.18); font-weight:600; }}
  section {{ background:var(--card); border:1px solid var(--line); border-radius:14px; padding:22px 24px; margin-bottom:20px; }}
  h2 {{ font-size:18px; margin:0 0 14px; display:flex; align-items:center; gap:8px; }}
  h2::before {{ content:""; width:4px; height:18px; background:var(--blue); border-radius:2px; }}
  .summary {{ color:var(--muted); }}
  .cards {{ display:grid; grid-template-columns:repeat(auto-fill,minmax(260px,1fr)); gap:14px; }}
  .rcard {{ border:1px solid var(--line); border-radius:12px; padding:14px 16px; border-left:4px solid var(--red); }}
  .rcard.mid {{ border-left-color:var(--orange); }}
  .rcard .t {{ font-weight:600; margin-bottom:4px; }}
  .rcard .d {{ font-size:13px; color:var(--muted); }}
  table {{ width:100%; border-collapse:collapse; font-size:13.5px; }}
  th,td {{ border:1px solid var(--line); padding:10px 12px; text-align:left; vertical-align:top; }}
  th {{ background:#1e40af; color:#fff; font-weight:600; }}
  tr:nth-child(even) td {{ background:#fafbfc; }}
  .badge {{ display:inline-block; padding:2px 9px; border-radius:999px; font-size:12px; font-weight:600; white-space:nowrap; }}
  .b-high {{ background:#fee4e2; color:var(--red); }}
  .b-mid {{ background:#fef0c7; color:#b54708; }}
  .b-low {{ background:#e0f2fe; color:#075985; }}
  .b-ok {{ background:#d1fadf; color:var(--green); }}
  .quote {{ background:#f8fafc; border:1px dashed var(--line); border-radius:8px; padding:8px 10px;
    color:#374151; font-size:12.5px; max-height:90px; overflow:auto; }}
  .comment {{ margin-top:8px; background:#fffbeb; border:1px solid #fde68a; border-left:4px solid var(--orange);
    border-radius:8px; padding:8px 10px; font-size:12.5px; color:#92400e; }}
  .comment b {{ color:#b45309; }}
  .emp {{ background:#f0fdf4; border:1px solid #bbf7d0; border-left:4px solid var(--green);
    border-radius:8px; padding:8px 10px; font-size:12.5px; color:#065f46; }}
  .emp b {{ color:#047857; }}
  .disclaimer {{ font-size:12.5px; color:var(--muted); border-top:1px dashed var(--line); padding-top:14px; }}
  .juris {{ font-size:13px; color:#075985; background:#e0f2fe; border-radius:8px; padding:10px 14px; margin-top:12px; }}
  .juris b {{ color:#075985; }}
  .juris .lr {{ display:block; margin-top:4px; color:#344054; font-size:12.5px; }}
</style>
</head>
<body>
<div class="wrap">
  <header class="{header_cls}">
    <h1>{title}</h1>
    <div class="meta">
      <div><span>文件名</span>{file_name}</div>
      <div><span>审查日期</span>{review_date}</div>
      <div><span>条款数</span>{clause_count}</div>
    </div>
    <div class="overall">总体风险评级：{overall_rating}</div>
  </header>

  <section>
    <h2>一、总体结论</h2>
    <p class="summary">{overall_summary}</p>
    {juris_block}
  </section>

  <section>
    <h2>二、重点风险摘要</h2>
    <div class="cards">
      {risk_cards}
    </div>
  </section>

  <section>
    <h2>三、逐条合规检查明细（按风险从高到低）</h2>
    <table>
      <thead>
        <tr>
          <th style="width:36px">#</th>
          <th style="width:190px">条款原文摘要</th>
          <th style="width:84px">检查维度</th>
          <th style="width:64px">结论</th>
          <th style="width:60px">风险</th>
          <th style="width:110px">法律依据</th>
          <th>建议改正内容</th>
          <th style="width:210px">给员工的建议</th>
        </tr>
      </thead>
      <tbody>
        {rows}
      </tbody>
    </table>
  </section>

  <section>
    <div class="disclaimer"><strong>免责声明：</strong>{disclaimer}</div>
  </section>
</div>
</body>
</html>"""


def esc(text):
    return html_mod.escape(str(text or ""))


def sort_rows(rows):
    """按风险级别从高到低排序（高>中>低>合规），同级别按原序号。"""
    return sorted(
        rows,
        key=lambda r: (RISK_ORDER.get(r.get("risk", "低"), 3), r.get("index", 0)),
    )


def build_jurisdiction(data):
    """构造地域信息块（国家/省/市 + 检索到的地方法规）。"""
    j = data.get("jurisdiction")
    if isinstance(j, dict):
        juris = " / ".join(x for x in [j.get("country"), j.get("province"), j.get("city")] if x) or "—"
    else:
        juris = str(j) if j else "—"
    lines = [f"📍 <b>适用地域（国家 / 省 / 市）：{esc(juris)}</b>"]
    if data.get("local_wage"):
        lines.append(esc(data["local_wage"]))
    for r in data.get("local_rules", []):
        lvl = esc(r.get("level", ""))
        topic = esc(r.get("topic", ""))
        val = esc(r.get("value", ""))
        src = esc(r.get("source", ""))
        dt = esc(r.get("date", ""))
        lines.append(f"【{lvl}】{topic}：{val}（来源：{src}，{dt}）")
    return '<div class="juris">' + "<br>".join(lines) + "</div>"


def render_rows(rows):
    out = []
    for r in rows:
        risk = r.get("risk", "低")
        badge = RISK_BADGE.get(risk, "b-low")
        is_high = r.get("is_high", risk == "高")
        comment = ""
        if is_high and r.get("compliant_example"):
            comment = (
                '<div class="comment"><b>💡 可直接替换的合规表述：</b><br>'
                + esc(r["compliant_example"]) + "</div>"
            )
        last = esc(r.get("suggestion", "")) + comment
        emp = esc(r.get("employee_advice", "")) or "—"
        out.append(
            "<tr>"
            f'<td>{esc(r.get("index",""))}</td>'
            f'<td><div class="quote">{esc(r.get("quote",""))}</div></td>'
            f'<td>{esc(r.get("dimension",""))}</td>'
            f'<td>{esc(r.get("conclusion",""))}</td>'
            f'<td><span class="badge {badge}">{esc(risk)}</span></td>'
            f'<td>{esc(r.get("law",""))}</td>'
            f"<td>{last}</td>"
            f'<td><div class="emp">{emp}</div></td>'
            "</tr>"
        )
    return "\n".join(out)


def render_html(data, out_path):
    rating = data.get("overall_rating", "")
    header_cls = "danger" if "高" in rating else ""
    cards = []
    for c in data.get("risk_cards", []):
        cls = "rcard" if c.get("level") != "中" else "rcard mid"
        cards.append(f'<div class="{cls}"><div class="t">{esc(c.get("title",""))}</div>'
                     f'<div class="d">{esc(c.get("desc",""))}</div></div>')
    html = HTML_TEMPLATE.format(
        title=esc(data.get("title", "劳动合同合规审查报告")),
        header_cls=header_cls,
        file_name=esc(data.get("file_name", "")),
        review_date=esc(data.get("review_date", "")),
        clause_count=esc(data.get("clause_count", "")),
        overall_rating=esc(rating),
        overall_summary=esc(data.get("overall_summary", "")),
        juris_block=build_jurisdiction(data),
        risk_cards="\n".join(cards),
        rows=render_rows(sort_rows(data.get("rows", []))),
        disclaimer=esc(data.get("disclaimer", DEFAULT_DISCLAIMER)).replace("\n", "<br>"),
    )
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)
    return out_path


def render_xlsx(data, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "审查明细"
    NCOL = 9

    thin = Side(style="thin", color="D0D5DD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="1E40AF")
    head_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=16, color="1E40AF")
    sub_font = Font(size=10, color="475467")
    juris_font = Font(size=10, color="075985", bold=True)
    concl_font = Font(size=10, color="344054")
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center")
    head_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    r = 1
    # 标题
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
    ws.cell(r, 1, data.get("title", "劳动合同合规审查报告")).font = title_font
    r += 1
    # 元信息行
    meta = (f"文件名：{data.get('file_name','')}    审查日期：{data.get('review_date','')}    "
            f"条款数：{data.get('clause_count','')}    总体风险：{data.get('overall_rating','')}")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
    ws.cell(r, 1, meta).font = sub_font
    r += 1
    # 适用地域（国家/省/市）
    j = data.get("jurisdiction")
    if isinstance(j, dict):
        juris = " / ".join(x for x in [j.get("country"), j.get("province"), j.get("city")] if x) or "—"
    else:
        juris = str(j) if j else "—"
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
    ws.cell(r, 1, f"适用地域（国家 / 省 / 市）：{juris}").font = juris_font
    r += 1
    # 当地最低工资
    if data.get("local_wage"):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
        c = ws.cell(r, 1, "📍 " + data["local_wage"])
        c.font = Font(size=10, color="075985")
        c.fill = PatternFill("solid", fgColor="E0F2FE")
        r += 1
    # 检索到的地方法规（层级变量）
    for lr in data.get("local_rules", []):
        lvl = lr.get("level", "")
        topic = lr.get("topic", "")
        val = lr.get("value", "")
        src = lr.get("source", "")
        dt = lr.get("date", "")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
        ws.cell(r, 1, f"【{lvl}】{topic}：{val}（来源：{src}，{dt}）").font = Font(size=10, color="344054")
        r += 1
    # 总体结论
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
    cc = ws.cell(r, 1, "总体结论：" + data.get("overall_summary", ""))
    cc.alignment = wrap
    cc.font = concl_font
    ws.row_dimensions[r].height = 46
    r += 1

    # 表头
    hdr = r
    headers = ["#", "条款原文摘要", "检查维度", "结论", "风险", "法律依据", "建议改正内容", "给员工的建议", "合规表述示例"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(hdr, ci, h)
        c.fill = head_fill
        c.font = head_font
        c.alignment = head_align
        c.border = border
    ws.row_dimensions[hdr].height = 30

    # 数据（按风险降序）
    rows = sort_rows(data.get("rows", []))
    for i, row_data in enumerate(rows):
        rr = hdr + 1 + i
        risk = row_data.get("risk", "低")
        values = [
            row_data.get("index", ""),
            row_data.get("quote", ""),
            row_data.get("dimension", ""),
            row_data.get("conclusion", ""),
            risk,
            row_data.get("law", ""),
            row_data.get("suggestion", ""),
            row_data.get("employee_advice", ""),
            row_data.get("compliant_example", ""),
        ]
        for ci, v in enumerate(values, 1):
            c = ws.cell(rr, ci, v)
            c.alignment = wrap
            c.border = border
            if ci in (1, 4, 5):
                c.alignment = center
        # 风险单元格分级配色
        rc = ws.cell(rr, 5)
        rc.fill = PatternFill("solid", fgColor=RISK_FILL.get(risk, "FFFFFF"))
        rc.font = Font(bold=True, color=RISK_FONT.get(risk, "000000"))
        # 给员工的建议列（第8列）浅绿底
        ecell = ws.cell(rr, 8)
        ecell.fill = PatternFill("solid", fgColor="F0FDF4")
        ecell.font = Font(color="065F46")
        # 行高自适应（按文本长度估算）
        h = max(48, min(200, (len(str(row_data.get("suggestion", ""))) +
                            len(str(row_data.get("employee_advice", "")))) // 2))
        ws.row_dimensions[rr].height = h

    last = hdr + len(rows)

    # 自动筛选 + 冻结首行（表头下方）
    if rows:
        ws.auto_filter.ref = f"A{hdr}:I{last}"
    ws.freeze_panes = f"A{hdr + 1}"

    # 免责声明
    drow = last + 2
    ws.merge_cells(start_row=drow, start_column=1, end_row=drow, end_column=NCOL)
    dc = ws.cell(drow, 1, "免责声明：" + data.get("disclaimer", DEFAULT_DISCLAIMER))
    dc.alignment = wrap
    dc.font = Font(size=9, color="667085")
    ws.row_dimensions[drow].height = 42

    # 列宽
    widths = [5, 32, 12, 9, 7, 16, 42, 38, 36]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    wb.save(out_path)
    return out_path


def main():
    if len(sys.argv) < 2:
        print("用法: python build_report.py <analysis.json> [输出目录]", file=sys.stderr)
        sys.exit(1)
    json_path = sys.argv[1]
    out_dir = sys.argv[2] if len(sys.argv) > 2 else os.path.dirname(os.path.abspath(json_path))
    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)

    base = os.path.join(out_dir, "劳动合同合规审查报告")
    html_path = base + ".html"
    xlsx_path = base + ".xlsx"
    render_html(data, html_path)
    render_xlsx(data, xlsx_path)
    print(html_path)
    print(xlsx_path)


if __name__ == "__main__":
    main()
