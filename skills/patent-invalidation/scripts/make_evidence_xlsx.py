#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
make_evidence_xlsx.py —— Excel 证据清单模板生成
====================================================

本脚本生成标准化的"证据清单与特征比对表" Excel 模板，含 8 列：
    编号 | 证据名称 | 来源 | 公开日 | 证据类型 | 与无效理由对应 | 形式合法性 | 备注

包含 6 行预填示例（每种证据类型一个）：
    1. 中国专利文献（CNIPA epub）
    2. 国外专利文献（Google Patents）
    3. 非专利文献（论文/标准）
    4. 使用公开（销售）
    5. 使用公开（展会）
    6. 使用公开（网络）

特性:
    - 冻结表头（首行 + 标题行）
    - 列宽自适应
    - 表头深色背景 + 白色加粗
    - 数据行斑马纹
    - 边框 + 自动换行
    - 公开日列日期格式 YYYY-MM-DD
    - 中文支持（UTF-8）

用法:
    python scripts/make_evidence_xlsx.py [--out <path.xlsx>]

依赖:
    pip install openpyxl
"""
import argparse
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误: 需要安装 openpyxl。请运行: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ── 列定义 ────────────────────────────────────────────────

COLUMNS = [
    {"key": "no", "title": "编号", "width": 8},
    {"key": "name", "title": "证据名称", "width": 35},
    {"key": "source", "title": "来源", "width": 35},
    {"key": "pubdate", "title": "公开日", "width": 14},
    {"key": "type", "title": "证据类型", "width": 15},
    {"key": "reasons", "title": "与无效理由对应", "width": 25},
    {"key": "form_legality", "title": "形式合法性", "width": 35},
    {"key": "remark", "title": "备注", "width": 25},
]

# 预填 6 行示例
SAMPLE_ROWS = [
    {
        "no": "1",
        "name": "CN118658342B《一种XX装置》（对比文件1）",
        "source": "国家知识产权局中国专利公布公告系统 epub.cnipa.gov.cn 专利单行本 PDF（下载日期 YYYY-MM-DD）",
        "pubdate": "2020-05-12",
        "type": "中国专利文献",
        "reasons": "新颖性(法22.2) / 创造性(法22.3)",
        "form_legality": "官方 PDF 公证复印件（可不做公证，国家局公开 PDF 即采信）",
        "remark": "用于主攻现有技术",
    },
    {
        "no": "2",
        "name": "US10234567B2《Method and apparatus for ...》（对比文件2）",
        "source": "Google Patents patents.google.com/patent/US10234567B2/en 附图页（下载日期 YYYY-MM-DD）",
        "pubdate": "2019-03-14",
        "type": "国外专利文献",
        "reasons": "创造性(法22.3) 三步法「区别特征来源」",
        "form_legality": "Google Patents 截图 + 网址公证；如已获 Espacenet PDF 则附公证复印件",
        "remark": "与证据 1 组合攻创造性",
    },
    {
        "no": "3",
        "name": "《XX 技术手册》第 X 章，ISBN 978-7-XXX-XXXXX-X",
        "source": "机械工业出版社，2013 年 6 月第 1 版",
        "pubdate": "2013-06-01",
        "type": "非专利文献（书籍）",
        "reasons": "公知常识（区别特征属常规手段）",
        "form_legality": "原件（含版权页 + ISBN）+ 公证复印件",
        "remark": "创造性论证的公知常识证据；口审辩论终结前可补",
    },
    {
        "no": "4",
        "name": "使用公开——XX 产品销售证据链",
        "source": "证据 4-1 销售合同 / 4-2 发票 / 4-3 物流单 / 4-4 银行流水 / 4-5 实物 / 4-6 检测报告 / 4-7 宣传册",
        "pubdate": "2014-08-15（合同签订日）",
        "type": "使用公开（销售）",
        "reasons": "新颖性(法22.2) / 创造性(法22.3)",
        "form_legality": "全链条公证：销售合同/发票/物流/银行流水均做原件公证；产品实物公证购买封存",
        "remark": "由 use_evidence_builder.py template sale 生成的 7 份证据合并为一组",
    },
    {
        "no": "5",
        "name": "使用公开——XX 展会证据链",
        "source": "证据 5-1 展会公告 / 5-2 展位图 / 5-3 参展合同 / 5-4 宣传册 / 5-5 现场视频 / 5-6 媒体报道",
        "pubdate": "2014-09-10（展会开始日）",
        "type": "使用公开（展会）",
        "reasons": "新颖性(法22.2) / 创造性(法22.3)",
        "form_legality": "全链条公证：现场视频建议公证员现场见证拍摄；网页截图 + URL + 时间戳公证",
        "remark": "由 use_evidence_builder.py template exhibition 生成",
    },
    {
        "no": "6",
        "name": "使用公开——XX 官网/电商页面",
        "source": "URL: https://www.example.com/product/xxx （Wayback Machine 历史快照）",
        "pubdate": "2014-11-20（页面发布时间）",
        "type": "使用公开（网络）",
        "reasons": "新颖性(法22.2) / 创造性(法22.3)",
        "form_legality": "**公证员现场访问** + 完整截图 + 时间戳 + 公证书；Wayback Machine 第三方印证",
        "remark": "网络公开类须严格按电子证据司法解释取证",
    },
]

# ── 样式定义 ─────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFFFF")
TITLE_FILL = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="FF1F4E78")
DATA_FONT = Font(name="微软雅黑", size=10)
ALT_FILL = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN = Side(border_style="thin", color="FFAAAAAA")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


# ── 生成函数 ─────────────────────────────────────────────

def make_workbook(rows=None) -> Workbook:
    """生成证据清单工作簿。

    rows: 数据行列表（每项为 dict，键同 COLUMNS.key）。为 None 时使用预填示例。
    v1.0.9: 支持 --content 填实模式（从 content.json 的 evidence_list 填入真实证据）。
    """
    data_rows = rows if rows is not None else SAMPLE_ROWS
    wb = Workbook()
    ws = wb.active
    ws.title = "证据清单"

    # 顶部标题（合并 8 列）
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    title = ws.cell(row=1, column=1, value="无效宣告请求书——证据清单与特征比对表")
    title.font = TITLE_FONT
    title.fill = TITLE_FILL
    title.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 32

    # 表头（第 2 行）
    for col_idx, col in enumerate(COLUMNS, 1):
        c = ws.cell(row=2, column=col_idx, value=col["title"])
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = ALIGN_CENTER
        c.border = BORDER
    ws.row_dimensions[2].height = 36

    # 数据行（从第 3 行开始）
    for row_idx, row_data in enumerate(data_rows, 3):
        for col_idx, col in enumerate(COLUMNS, 1):
            v = row_data.get(col["key"], "")
            c = ws.cell(row=row_idx, column=col_idx, value=v)
            c.font = DATA_FONT
            c.alignment = ALIGN_LEFT if col["key"] != "no" else ALIGN_CENTER
            c.border = BORDER
            if row_idx % 2 == 0:  # 斑马纹
                c.fill = ALT_FILL
            if col["key"] == "pubdate" and v:
                c.number_format = "yyyy-mm-dd"
        ws.row_dimensions[row_idx].height = 60

    # 列宽
    for col_idx, col in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col["width"]

    # 冻结表头（前 2 行）
    ws.freeze_panes = "A3"

    # ── 第二页：使用说明 ──
    ws2 = wb.create_sheet("使用说明")
    ws2.column_dimensions["A"].width = 100
    help_lines = [
        ("无效宣告请求书——证据清单 Excel 模板使用说明", True, TITLE_FONT, TITLE_FILL, 28),
        ("", False, None, None, 8),
        ("1. 顶部标题与表头（2 行）已冻结，滚动时始终可见。", False, DATA_FONT, None, 18),
        ("2. 数据从第 3 行开始填入，每行 = 一项证据（或一组证据链）。", False, DATA_FONT, None, 18),
        ("3. 8 列含义：", False, Font(name="微软雅黑", size=10, bold=True), None, 18),
        ("   - 编号：自定义字母数字编号（如 1, 2, 2.1, 证据 A）", False, DATA_FONT, None, 16),
        ("   - 证据名称：含公开号/书名/产品名 + 简要描述", False, DATA_FONT, None, 16),
        ("   - 来源：完整 URL / 数据库 / 出处 + 下载日期", False, DATA_FONT, None, 16),
        ("   - 公开日：YYYY-MM-DD（早于目标专利申请日/优先权日）", False, DATA_FONT, None, 16),
        ("   - 证据类型：中国专利 / 国外专利 / 非专利文献 / 使用公开 / 标准 / 授权档案", False, DATA_FONT, None, 16),
        ("   - 与无效理由对应：法22.2 / 法22.3 / 法26.3 / 法26.4 / 法33 / 细则23.2(旧20.2) / 法9.1 / 法5·25", False, DATA_FONT, None, 16),
        ("   - 形式合法性：公证 / 公证复印件 / 域外认证 / 译文 / 原件 / 版权页 等", False, DATA_FONT, None, 16),
        ("   - 备注：M5 特征映射表的附图依据、与对比文件的关系等", False, DATA_FONT, None, 16),
        ("", False, None, None, 8),
        ("4. 预填 6 行示例（每种证据类型一个），可保留作参考或删除后填入实际数据。", False, DATA_FONT, None, 18),
        ("", False, None, None, 8),
        ("5. 提交请求书时，将本表 + 6 行示例的\"形式合法性\"列（公证手续）一并附后。", False, DATA_FONT, None, 18),
        ("6. 一份证据可能对应多条无效理由（用 / 分隔）；一条理由可由多份证据印证。", False, DATA_FONT, None, 18),
        ("", False, None, None, 8),
        ("7. **硬门槛**：所有证据的\"公开日\"必须早于目标专利申请日（有优先权的，早于优先权日）。", False, Font(name="微软雅黑", size=10, bold=True, color="FFC00000"), None, 18),
        ("8. 详见 references/证据组合与证明标准.md", False, DATA_FONT, None, 16),
        ("9. 使用公开类证据链模板：python scripts/use_evidence_builder.py template <type> --out <path>", False, Font(name="微软雅黑", size=10, italic=True, color="FF666666"), None, 16),
    ]
    for i, (text, is_title, font, fill, height) in enumerate(help_lines, 1):
        c = ws2.cell(row=i, column=1, value=text)
        if font:
            c.font = font
        if fill:
            c.fill = fill
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws2.row_dimensions[i].height = height

    return wb


def main() -> int:
    ap = argparse.ArgumentParser(
        description="生成 Excel 证据清单（模板模式 / --content 填实模式）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # 模板模式（6 行预填示例）
  python make_evidence_xlsx.py --out 证据清单.xlsx
  # 填实模式（从 content.json 的 evidence_list 填入真实证据）
  python make_evidence_xlsx.py --content content.json --out 证据清单.xlsx
        """,
    )
    ap.add_argument(
        "--out",
        default=None,
        help="输出路径（默认写到 skill 根目录 assets/证据清单_模板.xlsx）",
    )
    ap.add_argument(
        "--content",
        default=None,
        help="v1.0.9 填实模式：从 content.json 读取 evidence_list 填入真实证据（与 make_invalidation_doc 的 content.json 同源）",
    )
    args = ap.parse_args()

    if args.out:
        out_path = args.out
    else:
        out_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            "assets",
            "证据清单_模板.xlsx",
        )
    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)

    rows = None
    mode_label = "模板（6 行预填示例）"
    if args.content:
        import json
        with open(args.content, encoding="utf-8") as f:
            data = json.load(f)
        ev_list = data.get("evidence_list", [])
        # content.json 的 evidence_list 键与 COLUMNS.key 完全对应
        rows = [
            {col["key"]: ev.get(col["key"], "") for col in COLUMNS}
            for ev in ev_list
        ]
        mode_label = f"填实（{len(rows)} 项证据，来自 {os.path.basename(args.content)}）"

    wb = make_workbook(rows=rows)
    wb.save(out_path)
    print(f"[OK] 已生成: {out_path}")
    print(f"     - 工作表 1: 证据清单（8 列 × {len(rows) if rows is not None else len(SAMPLE_ROWS)} 行数据 · {mode_label}）")
    print(f"     - 工作表 2: 使用说明")
    return 0


if __name__ == "__main__":
    sys.exit(main())
