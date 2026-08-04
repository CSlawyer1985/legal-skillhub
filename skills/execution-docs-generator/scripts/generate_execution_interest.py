#!/usr/bin/env python3
"""
执行标的计算表生成器 v2 — 完整利率数据库 + 双段利息

核心能力：
1. 内置完整利率数据库（2006-2026），涵盖央行基准利率 → LPR 全时段
2. 输出双 sheet：利率数据库 + 执行标的计算表（一般债务利息 + 加倍部分利息）
3. 三种利率模式：
   - fixed：判决书明确约定固定利率（如"年利率24%"）
   - LPR：判决书使用"LPR"，自动按 LPR 变化节点分段计算
   - benchmark：判决书使用"同期银行贷款利率"，自动按基准利率/LPR 分段
4. 所有计算列为 Excel 公式，可点击核验
5. 相邻同利率段自动合并
6. 输出含「利率数据库」+「执行标的计算表」两个 sheet

用法：
  # 固定利率模式
  python3 generate_execution_interest.py out.xlsx \
    --principal 100000 \
    --general-start 2024-01-01 --general-end 2024-12-31 --general-rate 3.45 \
    --doubled-start 2025-01-01 --doubled-end 2026-07-16 \
    --submitter "张三" --case-info "(2024)苏0213民初123号"

  # LPR 模式（自动分段）
  python3 generate_execution_interest.py out.xlsx \
    --principal 100000 \
    --general-start 2024-01-01 --general-end 2026-07-16 --general-rate-type LPR \
    --doubled-start 2024-01-16 --doubled-end 2026-07-16 \
    --submitter "张三"

  # 同期银行贷款利率模式（自动分段）
  python3 generate_execution_interest.py out.xlsx \
    --principal 100000 \
    --general-start 2022-01-01 --general-end 2026-07-16 --general-rate-type benchmark \
    --doubled-start 2022-01-16 --doubled-end 2026-07-16 \
    --submitter "张三"
"""

import argparse
from datetime import date, timedelta, datetime
from dataclasses import dataclass
from typing import List, Optional, Tuple


# ======================== 利率数据库（完整内置，2006-2026） ========================

@dataclass
class RateRecord:
    effective_date: date
    rate_6m: float       # 六个月以内（含六个月）
    rate_6m_1y: float    # 六个月至一年（含一年）
    rate_1_3y: float     # 一至三年（含三年）
    rate_3_5y: float     # 三至五年（含五年）
    rate_5y_plus: float  # 五年以上


RATE_DATABASE: List[RateRecord] = [
    # === 央行贷款基准利率时期 ===
    RateRecord(date(2006, 8, 19),  5.58, 6.12, 6.30, 6.48, 6.84),
    RateRecord(date(2007, 3, 18),  5.67, 6.39, 6.57, 6.75, 7.11),
    RateRecord(date(2007, 5, 19),  5.85, 6.57, 6.75, 6.93, 7.20),
    RateRecord(date(2007, 7, 21),  6.03, 6.84, 7.02, 7.20, 7.38),
    RateRecord(date(2007, 8, 22),  6.21, 7.02, 7.20, 7.38, 7.56),
    RateRecord(date(2007, 9, 15),  6.48, 7.29, 7.47, 7.65, 7.83),
    RateRecord(date(2007, 12, 21), 6.57, 7.47, 7.56, 7.74, 7.83),
    RateRecord(date(2008, 9, 16),  6.21, 7.20, 7.29, 7.56, 7.74),
    RateRecord(date(2008, 10, 9),  6.12, 6.93, 7.02, 7.29, 7.47),
    RateRecord(date(2008, 10, 30), 6.03, 6.66, 6.75, 7.02, 7.20),
    RateRecord(date(2008, 11, 27), 5.04, 5.58, 5.67, 5.94, 6.12),
    RateRecord(date(2008, 12, 23), 4.86, 5.31, 5.40, 5.76, 5.94),
    RateRecord(date(2010, 10, 20), 5.10, 5.56, 5.60, 5.96, 6.14),
    RateRecord(date(2010, 12, 26), 5.35, 5.81, 5.85, 6.22, 6.40),
    RateRecord(date(2011, 2, 9),   5.60, 6.06, 6.10, 6.45, 6.60),
    RateRecord(date(2011, 4, 6),   5.85, 6.31, 6.40, 6.65, 6.80),
    RateRecord(date(2011, 7, 7),   6.10, 6.56, 6.65, 6.90, 7.05),
    RateRecord(date(2012, 6, 8),   5.85, 6.31, 6.40, 6.65, 6.80),
    RateRecord(date(2012, 7, 6),   5.60, 6.00, 6.15, 6.40, 6.55),
    RateRecord(date(2014, 11, 22), 5.60, 5.60, 6.00, 6.00, 6.15),
    RateRecord(date(2015, 3, 1),   5.35, 5.35, 5.75, 5.75, 5.90),
    RateRecord(date(2015, 5, 11),  5.10, 5.10, 5.50, 5.50, 5.65),
    RateRecord(date(2015, 6, 28),  4.85, 4.85, 5.25, 5.25, 5.40),
    RateRecord(date(2015, 8, 26),  4.60, 4.60, 5.00, 5.00, 5.15),
    RateRecord(date(2015, 10, 24), 4.35, 4.35, 4.75, 4.75, 4.90),
    # === LPR 时期（2019.8 起） ===
    RateRecord(date(2019, 8, 20),  4.25, 4.25, 4.25, 4.25, 4.85),
    RateRecord(date(2019, 9, 20),  4.20, 4.20, 4.20, 4.20, 4.85),
    RateRecord(date(2019, 10, 21), 4.20, 4.20, 4.20, 4.20, 4.85),
    RateRecord(date(2019, 11, 20), 4.15, 4.15, 4.15, 4.15, 4.80),
    RateRecord(date(2019, 12, 20), 4.15, 4.15, 4.15, 4.15, 4.80),
    RateRecord(date(2020, 1, 20),  4.15, 4.15, 4.15, 4.15, 4.80),
    RateRecord(date(2020, 2, 20),  4.05, 4.05, 4.05, 4.05, 4.75),
    RateRecord(date(2020, 3, 20),  4.05, 4.05, 4.05, 4.05, 4.75),
    RateRecord(date(2020, 4, 20),  3.85, 3.85, 3.85, 3.85, 4.65),
    RateRecord(date(2021, 12, 20), 3.80, 3.80, 3.80, 3.80, 4.65),
    RateRecord(date(2022, 1, 20),  3.70, 3.70, 3.70, 3.70, 4.60),
    RateRecord(date(2022, 5, 20),  3.70, 3.70, 3.70, 3.70, 4.45),
    RateRecord(date(2022, 8, 22),  3.65, 3.65, 3.65, 3.65, 4.30),
    RateRecord(date(2023, 6, 20),  3.55, 3.55, 3.55, 3.55, 4.20),
    RateRecord(date(2023, 8, 21),  3.45, 3.45, 3.45, 3.45, 4.20),
    RateRecord(date(2024, 2, 20),  3.45, 3.45, 3.45, 3.45, 3.95),
    RateRecord(date(2024, 7, 22),  3.35, 3.35, 3.35, 3.35, 3.85),
    RateRecord(date(2024, 10, 21), 3.10, 3.10, 3.10, 3.10, 3.60),
    RateRecord(date(2024, 11, 20), 3.10, 3.10, 3.10, 3.10, 3.60),
    RateRecord(date(2024, 12, 20), 3.10, 3.10, 3.10, 3.10, 3.60),
    RateRecord(date(2025, 1, 20),  3.10, 3.10, 3.10, 3.10, 3.60),
    RateRecord(date(2025, 2, 20),  3.10, 3.10, 3.10, 3.10, 3.60),
    RateRecord(date(2025, 3, 20),  3.10, 3.10, 3.10, 3.10, 3.60),
    RateRecord(date(2025, 4, 21),  3.10, 3.10, 3.10, 3.10, 3.60),
    RateRecord(date(2025, 5, 20),  3.00, 3.00, 3.00, 3.00, 3.50),
    RateRecord(date(2025, 6, 20),  3.00, 3.00, 3.00, 3.00, 3.50),
    RateRecord(date(2025, 7, 21),  3.00, 3.00, 3.00, 3.00, 3.50),
    RateRecord(date(2025, 8, 20),  3.00, 3.00, 3.00, 3.00, 3.50),
    RateRecord(date(2025, 9, 22),  3.00, 3.00, 3.00, 3.00, 3.50),
    RateRecord(date(2025, 10, 20), 3.00, 3.00, 3.00, 3.00, 3.50),
    RateRecord(date(2025, 11, 20), 3.00, 3.00, 3.00, 3.00, 3.50),
    RateRecord(date(2025, 12, 22), 3.00, 3.00, 3.00, 3.00, 3.50),
    RateRecord(date(2026, 1, 20),  3.00, 3.00, 3.00, 3.00, 3.50),
    RateRecord(date(2026, 2, 20),  3.00, 3.00, 3.00, 3.00, 3.50),
    RateRecord(date(2026, 3, 20),  3.00, 3.00, 3.00, 3.00, 3.50),
    RateRecord(date(2026, 4, 20),  3.00, 3.00, 3.00, 3.00, 3.50),
    RateRecord(date(2026, 5, 20),  3.00, 3.00, 3.00, 3.00, 3.50),
    RateRecord(date(2026, 6, 22),  3.00, 3.00, 3.00, 3.00, 3.50),
]

# 确保按生效日期排序
RATE_DATABASE.sort(key=lambda r: r.effective_date)

# 档位属性
_RATE_ATTRS = ["rate_6m", "rate_6m_1y", "rate_1_3y", "rate_3_5y", "rate_5y_plus"]
_RATE_COL_NAMES = ["六个月以内", "六个月至一年", "一至三年", "三至五年", "五年以上"]
_RATE_TERM_MAP = {"6M": 0, "1Y": 1, "1-3Y": 2, "3-5Y": 3, "5Y": 4}

# LPR 制度起始日期
LPR_START_DATE = date(2019, 8, 20)


# ======================== 利率选择 / 分段逻辑 ========================

def get_rate(rec: RateRecord, term: str) -> float:
    """获取指定档位的利率"""
    idx = _RATE_TERM_MAP.get(term, 1)
    return getattr(rec, _RATE_ATTRS[idx])


def auto_select_term(start: date, end: date) -> str:
    """根据期间长度自动选择利率档位"""
    years = (end - start).days / 365.25
    if years < 0.5:
        return "6M"
    elif years < 1:
        return "1Y"
    elif years < 3:
        return "1-3Y"
    elif years < 5:
        return "3-5Y"
    else:
        return "5Y"


def applicable_rate(db: List[RateRecord], d: date, term: str) -> float:
    """返回指定日期适用的利率"""
    rec = db[0]
    for r in db:
        if r.effective_date <= d:
            rec = r
        else:
            break
    return get_rate(rec, term)


def split_periods(
    start: date,
    end: date,
    db: List[RateRecord],
    term: str = "1Y",
    fixed_rate: Optional[float] = None,
) -> List[Tuple[date, date, float]]:
    """
    按利率变化节点拆分期间。

    - fixed_rate 有值：不分段，直接返回一段
    - fixed_rate 为 None：按利率数据库变化节点自动分段

    返回 [(分段起始日, 分段截止日, 适用利率), ...]
    利率为年利率百分比（如 3.45 表示 3.45%）
    """
    if fixed_rate is not None and fixed_rate > 0:
        return [(start, end, fixed_rate)]

    periods: List[Tuple[date, date, float]] = []
    current = start

    for r in db:
        if r.effective_date <= current:
            continue
        if r.effective_date > end:
            rate = applicable_rate(db, current, term)
            periods.append((current, end, rate))
            current = end
            break
        period_end = r.effective_date - timedelta(days=1)
        if period_end >= current:
            rate = applicable_rate(db, current, term)
            periods.append((current, min(period_end, end), rate))
            current = r.effective_date

    if current < end:
        periods.append((current, end, applicable_rate(db, current, term)))

    # 合并相邻同利率段
    merged: List[Tuple[date, date, float]] = []
    for ps, pe, pr in periods:
        if merged and merged[-1][2] == pr:
            merged[-1] = (merged[-1][0], pe, pr)
        else:
            merged.append((ps, pe, pr))
    return merged


# ======================== Excel 生成（双 sheet） ========================

try:
    from openpyxl import Workbook
    from openpyxl.workbook.properties import CalcProperties
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
TITLE_FONT = Font(name="宋体", size=14, bold=True)
SECTION_FONT = Font(name="宋体", size=12, bold=True)
HEADER_FONT = Font(name="宋体", size=11, bold=True)
BODY_FONT = Font(name="宋体", size=11)
BOLD_FONT = Font(name="宋体", size=11, bold=True)
TINY_FONT = Font(name="宋体", size=9, color="999999")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _set_cell(ws, row, col, value, font=None, alignment=None, fill=None, number_format=None, border=True):
    """设置单元格格式"""
    cell = ws.cell(row=row, column=col, value=value)
    if border:
        cell.border = THIN_BORDER
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if fill:
        cell.fill = fill
    if number_format:
        cell.number_format = number_format
    return cell


def _build_rate_db_sheet(ws):
    """构建利率数据库 sheet（完整历史数据）"""
    title_font = Font(name="宋体", size=11, bold=True)
    data_font = Font(name="宋体", size=10)
    header_font = Font(name="宋体", size=10, bold=True)
    center = Alignment(horizontal="center", vertical="center")

    widths = {"A": 14, "B": 16, "C": 16, "D": 14, "E": 14, "F": 14}
    for c, w in widths.items():
        ws.column_dimensions[c].width = w

    # 来源链接
    ws.cell(row=1, column=1, value="人民银行公布链接").font = Font(name="宋体", size=9, color="666666")

    # 档位断点行
    for ci, (val, col) in enumerate([(0, 2), (0.5, 3), (1, 4), (3, 5), (5, 6)], 1):
        ws.cell(row=2, column=col, value=val).font = data_font

    # 表头
    headers = ["调整时间", "六个月以内\n（含六个月）", "六个月至一年\n（含一年）",
               "一至三年\n（含三年）", "三至五年\n（含五年）", "五年以上"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = header_font
        c.alignment = center
        c.border = THIN_BORDER

    # 数据
    for i, rec in enumerate(RATE_DATABASE):
        r = i + 4
        ws.cell(row=r, column=1, value=rec.effective_date).font = data_font
        ws.cell(row=r, column=1).alignment = center
        ws.cell(row=r, column=1).number_format = 'yyyy-mm-dd'
        ws.cell(row=r, column=1).border = THIN_BORDER
        for ci, attr in enumerate(_RATE_ATTRS, 2):
            c = ws.cell(row=r, column=ci, value=getattr(rec, attr))
            c.font = data_font
            c.alignment = center
            c.number_format = '0.00'
            c.border = THIN_BORDER


def _build_execution_sheet(
    ws,
    principal: float,
    general_periods: List[Tuple[date, date, float]],
    general_rate_type: str,
    doubled_start: date,
    doubled_end: date,
    doubled_daily_rate: float,
    submitter: str,
    case_info: str,
    rate_term: str,
    litigation_fee: float = 0,
):
    """构建执行标的计算表 sheet（一般债务利息 + 加倍部分利息）"""
    # 列宽
    for cw in [("A", 6), ("B", 14), ("C", 14), ("D", 14), ("E", 14), ("F", 16)]:
        ws.column_dimensions[cw[0]].width = cw[1]

    row = 1

    # ---- 标题 ----
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    _set_cell(ws, row, 1, "执行标的计算表", font=TITLE_FONT, alignment=CENTER, border=False)
    row += 1

    if case_info:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        _set_cell(ws, row, 1, case_info, font=BODY_FONT, alignment=CENTER, border=False)
        row += 1

    row += 1  # 空行

    # ---- 汇总行 ----
    sum_headers = ["", "本金（元）", "一般债务利息（元）", "加倍部分利息（元）", "合计（元）", ""]
    for ci, h in enumerate(sum_headers):
        _set_cell(ws, row, ci + 1, h, font=HEADER_FONT, alignment=CENTER, fill=HEADER_FILL)
    row += 1
    sum_row = row  # 记录汇总行号

    _set_cell(ws, row, 1, "", font=BODY_FONT)
    _set_cell(ws, row, 2, principal, font=BODY_FONT, alignment=RIGHT, number_format="#,##0.00")

    # 公式占位，后面更新
    _set_cell(ws, row, 3, 0, font=BODY_FONT, alignment=RIGHT, number_format="#,##0.00")
    _set_cell(ws, row, 4, 0, font=BODY_FONT, alignment=RIGHT, number_format="#,##0.00")
    _set_cell(ws, row, 5, 0, font=HEADER_FONT, alignment=RIGHT, number_format="#,##0.00")
    row += 2

    # ============================================================
    # 一、一般债务利息
    # ============================================================
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    _set_cell(ws, row, 1, "一、一般债务利息", font=SECTION_FONT, alignment=LEFT_WRAP, border=False)
    row += 1

    # 利率模式说明
    if general_rate_type == "fixed":
        if general_periods:
            rate_note = f"判决书约定固定年利率：{general_periods[0][2]}%"
        else:
            rate_note = "固定利率"
    elif general_rate_type == "benchmark":
        rate_note = f"判词采用「同期银行贷款利率」，档位：{rate_term}，按利率变化节点自动分段"
    else:
        rate_note = f"判词采用「LPR」，档位：{rate_term}，按 LPR 变化节点自动分段"

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    _set_cell(ws, row, 1, rate_note, font=Font(name="宋体", size=9, color="666666"), border=False)
    row += 1

    # 表头
    detail_headers = ["序号", "本金（元）", "起算日期", "截止日期", "年利率", "利息（元）"]
    for ci, h in enumerate(detail_headers):
        _set_cell(ws, row, ci + 1, h, font=HEADER_FONT, alignment=CENTER, fill=HEADER_FILL)
    row += 1

    general_data_start = row
    for idx, (s, e, r) in enumerate(general_periods, 1):
        _set_cell(ws, row, 1, str(idx), font=BODY_FONT, alignment=CENTER)
        _set_cell(ws, row, 2, principal, font=BODY_FONT, alignment=RIGHT, number_format="#,##0.00")
        _set_cell(ws, row, 3, s, font=BODY_FONT, alignment=CENTER)
        _set_cell(ws, row, 4, e, font=BODY_FONT, alignment=CENTER)
        _set_cell(ws, row, 5, r / 100, font=BODY_FONT, alignment=CENTER, number_format="0.000%")

        # 利息公式: =本金 * (截止日-起算日+1) * 年利率 / 360
        interest_formula = f"=B{row}*(D{row}-C{row}+1)*E{row}/360"
        _set_cell(ws, row, 6, interest_formula, font=BODY_FONT, alignment=RIGHT, number_format="#,##0.00")
        row += 1

    general_data_end = row - 1

    # 小计
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    _set_cell(ws, row, 1, "小计", font=HEADER_FONT, alignment=RIGHT)
    general_sum_formula = f"=SUM(F{general_data_start}:F{general_data_end})" if general_periods else "0"
    _set_cell(ws, row, 6, general_sum_formula, font=HEADER_FONT, alignment=RIGHT, number_format="#,##0.00")
    general_total_row = row
    row += 2

    # ============================================================
    # 二、加倍部分债务利息（《民诉法》第264条）
    # ============================================================
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    _set_cell(ws, row, 1, "二、加倍部分债务利息（《民诉法》第264条）", font=SECTION_FONT, alignment=LEFT_WRAP, border=False)
    row += 1

    # 说明
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    _set_cell(ws, row, 1, f"日利率：{doubled_daily_rate:.4f}%（万分之{doubled_daily_rate*10000:.2f}）",
              font=Font(name="宋体", size=9, color="666666"), border=False)
    row += 1

    doubled_headers = ["序号", "本金（元）", "实际起算日", "截止日", "日利率", "利息（元）"]
    for ci, h in enumerate(doubled_headers):
        _set_cell(ws, row, ci + 1, h, font=HEADER_FONT, alignment=CENTER, fill=HEADER_FILL)
    row += 1

    doubled_data_start = row
    _set_cell(ws, row, 1, "1", font=BODY_FONT, alignment=CENTER)
    _set_cell(ws, row, 2, principal, font=BODY_FONT, alignment=RIGHT, number_format="#,##0.00")
    _set_cell(ws, row, 3, doubled_start, font=BODY_FONT, alignment=CENTER)
    _set_cell(ws, row, 4, doubled_end, font=BODY_FONT, alignment=CENTER)
    _set_cell(ws, row, 5, doubled_daily_rate, font=BODY_FONT, alignment=CENTER, number_format="0.0000%")

    # 加倍利息公式: =本金 * (截止日-起算日+1) * 日利率
    doubled_interest_formula = f"=B{row}*(D{row}-C{row}+1)*E{row}"
    _set_cell(ws, row, 6, doubled_interest_formula, font=BODY_FONT, alignment=RIGHT, number_format="#,##0.00")
    row += 1
    doubled_data_end = row - 1

    # 小计
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    _set_cell(ws, row, 1, "小计", font=HEADER_FONT, alignment=RIGHT)
    doubled_sum_formula = f"=SUM(F{doubled_data_start}:F{doubled_data_end})"
    _set_cell(ws, row, 6, doubled_sum_formula, font=HEADER_FONT, alignment=RIGHT, number_format="#,##0.00")
    doubled_total_row = row
    row += 2

    # ============================================================
    # 三、合计
    # ============================================================
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    _set_cell(ws, row, 1, "合计（本金 + 一般债务利息 + 加倍部分债务利息）", font=HEADER_FONT, alignment=RIGHT)
    total_formula = f"=B{sum_row}+F{general_total_row}+F{doubled_total_row}"
    _set_cell(ws, row, 6, total_formula, font=SECTION_FONT, alignment=RIGHT, number_format="#,##0.00")
    total_row = row
    row += 2

    # ---- 更新汇总行公式 ----
    _set_cell(ws, sum_row, 3, f"=F{general_total_row}", font=BODY_FONT, alignment=RIGHT, number_format="#,##0.00")
    _set_cell(ws, sum_row, 4, f"=F{doubled_total_row}", font=BODY_FONT, alignment=RIGHT, number_format="#,##0.00")
    _set_cell(ws, sum_row, 5, f"=F{total_row}", font=HEADER_FONT, alignment=RIGHT, number_format="#,##0.00")

    # ---- 提交人签字 ----
    row += 1
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
    _set_cell(ws, row, 3, f"提交人：{submitter}", font=BODY_FONT, alignment=RIGHT, border=False)
    row += 1
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
    _set_cell(ws, row, 3, f"日期：{datetime.now().strftime('%Y年%m月%d日')}", font=BODY_FONT, alignment=RIGHT, border=False)


def generate_excel(
    output_path: str,
    principal: float,
    general_periods: List[Tuple[date, date, float]],
    general_rate_type: str,
    doubled_start: date,
    doubled_end: date,
    doubled_daily_rate: float,
    submitter: str,
    case_info: str = "",
    rate_term: str = "1Y",
):
    """生成执行标的计算表 Excel（双 sheet）"""
    if not _HAS_OPENPYXL:
        print("错误：需要 openpyxl 库。pip install openpyxl")
        import sys
        sys.exit(1)

    wb = Workbook()
    wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True)

    # Sheet 1: 执行标的计算表（默认可见）
    ws_exec = wb.active
    ws_exec.title = "执行标的计算表"
    _build_execution_sheet(
        ws_exec, principal, general_periods, general_rate_type,
        doubled_start, doubled_end, doubled_daily_rate,
        submitter, case_info, rate_term,
    )

    # Sheet 2: 利率数据库
    ws_db = wb.create_sheet("利率数据库")
    _build_rate_db_sheet(ws_db)

    wb.save(output_path)
    return output_path


# ======================== CLI ========================

def main():
    parser = argparse.ArgumentParser(
        description="执行标的计算表生成器 v2 — 完整利率数据库 + 双段利息",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("output", help="输出 .xlsx 路径")
    parser.add_argument("--principal", type=float, required=True, help="本金（元）")

    # 一般债务利息参数
    parser.add_argument("--general-start", required=True, help="一般债务利息起算日 YYYY-MM-DD")
    parser.add_argument("--general-end", required=True, help="一般债务利息截止日 YYYY-MM-DD")
    parser.add_argument(
        "--general-rate", type=float,
        help="一般债务利息年利率（%，如 3.45）。提供即使用固定利率模式"
    )
    parser.add_argument(
        "--general-rate-type", default="LPR",
        choices=["fixed", "LPR", "benchmark"],
        help="利率模式：fixed（固定利率）、LPR（按 LPR 分段）、benchmark（按同期银行贷款利率分段）。"
             " 当 --general-rate 有值时自动为 fixed 模式，忽略此参数。默认 LPR"
    )
    parser.add_argument(
        "--rate-term", default="1Y",
        choices=["6M", "1Y", "1-3Y", "3-5Y", "5Y"],
        help="利率档位（LPR/benchmark 模式）。默认 1Y（一年期）"
    )

    # 加倍部分利息参数
    parser.add_argument("--doubled-start", required=True, help="加倍部分利息起算日 YYYY-MM-DD")
    parser.add_argument("--doubled-end", required=True, help="加倍部分利息截止日 YYYY-MM-DD")
    parser.add_argument("--doubled-rate", type=float, default=0.000175,
                        help="加倍部分日利率（默认 0.000175 = 万分之 1.75）")

    parser.add_argument("--submitter", default="", help="提交人姓名")
    parser.add_argument("--case-info", default="", help="案号等信息，显示在标题下方")

    args = parser.parse_args()

    general_start = date.fromisoformat(args.general_start)
    general_end = date.fromisoformat(args.general_end)
    doubled_start = date.fromisoformat(args.doubled_start)
    doubled_end = date.fromisoformat(args.doubled_end)

    # 确定利率模式和期间
    if args.general_rate is not None and args.general_rate > 0:
        rate_type = "fixed"
        term = "1Y"  # 固定利率不用档位
        general_periods = [(general_start, general_end, args.general_rate)]
    else:
        rate_type = args.general_rate_type
        term = args.rate_term

        # 如果用户没指定档位，自动选择
        if args.rate_term == "1Y" and rate_type != "fixed":
            term = auto_select_term(general_start, general_end)

        general_periods = split_periods(
            general_start, general_end, RATE_DATABASE, term, fixed_rate=None
        )

    generate_excel(
        output_path=args.output,
        principal=args.principal,
        general_periods=general_periods,
        general_rate_type=rate_type,
        doubled_start=doubled_start,
        doubled_end=doubled_end,
        doubled_daily_rate=args.doubled_rate,
        submitter=args.submitter,
        case_info=args.case_info,
        rate_term=term,
    )

    # 输出摘要
    if rate_type == "fixed":
        tag = f"固定利率 {args.general_rate}%"
    elif rate_type == "benchmark":
        tag = f"同期银行贷款利率（{term} 档），按利率节点自动分段"
    else:
        tag = f"LPR（{term} 档），按利率变化自动分段"

    print(f"  本金 {args.principal:,.2f} 元")
    print(f"  一般债务利息：{args.general_start} ~ {args.general_end}，{tag}，共 {len(general_periods)} 段")
    for ps, pe, pr in general_periods:
        days = (pe - ps).days + 1
        print(f"    {ps} ~ {pe}：{pr}%，{days} 天")
    print(f"  加倍部分利息：{args.doubled_start} ~ {args.doubled_end}，日利率 {args.doubled_rate:.4f}%")
    print(f"\n✅ 已生成：{args.output}")
    print(f"   含「执行标的计算表」+「利率数据库」两个 sheet")


if __name__ == "__main__":
    main()
